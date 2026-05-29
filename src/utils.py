"""
Utility functions for data loading and result persistence.

Functions
---------
load_representation  — load a preprocessed taxa .npz and return train/test arrays
save_rep_to_excel    — write all result sections for one representation into a single sheet
"""

import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from config import RESULTS_XLSX


def load_representation(taxa_path):
    """Load a preprocessed taxa .npz file and return train/test splits.

    Args:
        taxa_path : str — path to .npz file (e.g. preprocessed_binary.npz)

    Returns:
        X_train : float32 array — training images
        X_test  : float32 array — test images
        y_train : float32 array — training labels
        y_test  : float32 array — test labels
    """
    npz     = np.load(taxa_path)
    X_train = npz['X_train'].astype('float32')
    X_test  = npz['X_test'].astype('float32')
    y_train = npz['y_train'].astype('float32')
    y_test  = npz['y_test'].astype('float32')
    return X_train, X_test, y_train, y_test


def save_rep_to_excel(rep_name, gs_df, cv_metrics, test_metrics,
                      perm_df=None, grad_df=None, grouped_perm_df=None,
                      results_path=None):
    """Write all result sections for one representation into a single sheet.

    Each section is preceded by a bold title row, then the DataFrame header
    and data. Two blank rows separate consecutive sections.  One sheet per
    representation keeps the workbook tidy (e.g. 'binary_io', 'binary_meta').

    Sheet layout example (rep_name = 'binary_meta'):
      Row  1 : ── GRID SEARCH ──                   (bold)
      Row  2 : column headers
      Rows 3–14 : grid-search rows
      Rows 15–16 : blank
      Row 17 : ── K-FOLD CV METRICS ──             (bold)
      …
      ── PERMUTATION IMPORTANCE ──
      ── GRADIENT IMPORTANCE ──
      ── GROUPED PERMUTATION IMPORTANCE ──

    Args:
        rep_name        : str        — sheet name (e.g. 'binary_io', 'normalized_meta')
        gs_df           : DataFrame  — from run_grid_search
        cv_metrics      : list[dict] — from run_kfold_cv
        test_metrics    : dict       — from run_final_model_*
        perm_df         : DataFrame or None — individual permutation importance
        grad_df         : DataFrame or None — gradient importance
        grouped_perm_df : DataFrame or None — grouped permutation importance
        results_path    : str or None — override default RESULTS_XLSX
    """
    path = results_path or RESULTS_XLSX
    os.makedirs(os.path.dirname(path), exist_ok=True)

    cv_df = pd.DataFrame(cv_metrics)
    cv_df.insert(0, 'fold', range(1, len(cv_metrics) + 1))
    test_df = pd.DataFrame([test_metrics])

    sections = [
        ('GRID SEARCH',       gs_df),
        ('K-FOLD CV METRICS', cv_df),
        ('TEST SET METRICS',  test_df),
    ]
    if perm_df is not None:
        sections.append(('PERMUTATION IMPORTANCE', perm_df))
    if grad_df is not None:
        sections.append(('GRADIENT IMPORTANCE', grad_df))
    if grouped_perm_df is not None:
        sections.append(('GROUPED PERMUTATION IMPORTANCE', grouped_perm_df))

    # Open existing workbook or create a fresh one
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        # Workbook() creates a default 'Sheet' — remove it
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    # Drop any prior version of this rep's sheet so we start fresh
    if rep_name in wb.sheetnames:
        del wb[rep_name]
    ws = wb.create_sheet(rep_name)

    # Write each section: bold title row, then DataFrame (header + data), then 2 blanks
    row = 1   # openpyxl is 1-indexed
    for title, df in sections:
        # Bold title row
        ws.cell(row=row, column=1, value=f'── {title} ──').font = Font(bold=True)
        row += 1
        # Header row
        for c, name in enumerate(df.columns, start=1):
            ws.cell(row=row, column=c, value=name)
        row += 1
        # Data rows (convert numpy scalars to native Python so openpyxl is happy)
        for _, data_row in df.iterrows():
            for c, val in enumerate(data_row, start=1):
                if hasattr(val, 'item'):
                    val = val.item()
                ws.cell(row=row, column=c, value=val)
            row += 1
        # Two blank rows between sections
        row += 2

    # openpyxl requires at least one sheet to save
    if not wb.sheetnames:
        wb.create_sheet('_placeholder')
    wb.save(path)

    print(f"  Saved -> {path}  (sheet: {rep_name})")
